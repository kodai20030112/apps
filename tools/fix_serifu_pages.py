#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台本ページ修正ツール（おぼえりふ用）

AI（Claude等）が作った台本Excelの「ページ」列は、見開きPDFの境目で
1ページずれることがある。このスクリプトは元のPDFを直接読み、
各セリフが「実際に印刷されているページ」を割り出して、ページ列を振り直す。

使い方:
    python3 fix_serifu_pages.py <台本.pdf> <AIが作った.xlsx> [出力.xlsx]

前提:
    - 縦書き（右→左）の日本語台本PDF
    - 各ページ（または見開きの各面）の下に「-12-」のようなページ番号が印刷されている
    - Excelは「番号／ページ／登場人物／セリフ」の並び（おぼえりふテンプレート）

必要ライブラリ: pymupdf, openpyxl  （.command が自動で入れます）
"""
import sys, re, unicodedata

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("PyMuPDF が入っていません。`pip3 install pymupdf` を実行してください。")
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が入っていません。`pip3 install openpyxl` を実行してください。")


def norm(s):
    """比較用に正規化（全半角を統一し、空白を除く）"""
    if s is None:
        return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(s)))


def clean_word(s):
    return s.replace("　", "").replace(" ", "").strip()


PAGENUM_RE = re.compile(r"^[\-‐‒–—－\s]*(\d{1,3})[\-‐‒–—－\s]*$")


def page_markers(words, H):
    """ページ下部にある『-N-』のようなページ番号トークンを拾う → [(番号, x中心)]"""
    res = []
    for w in words:
        c = clean_word(w[4])
        m = PAGENUM_RE.match(c)
        if m and ("-" in c or "‐" in c or "－" in c) and w[3] > H * 0.85:
            n = int(m.group(1))
            if 1 <= n <= 999:
                res.append((n, (w[0] + w[2]) / 2))
    return res


def reconstruct(words):
    """縦書き（右→左・上から下）の読み順にテキストを復元する"""
    items = []
    for w in words:
        t = clean_word(w[4])
        if not t:
            continue
        items.append(((w[0] + w[2]) / 2, w[1], t))
    if not items:
        return ""
    items.sort(key=lambda a: -a[0])            # 列は右から左へ
    cols, cur, curx = [], [], None
    for xc, y, t in items:
        if curx is None or abs(xc - curx) <= 9:  # 同じ列（x がほぼ同じ）
            cur.append((y, t))
            curx = xc if curx is None else (curx + xc) / 2
        else:
            cols.append(cur); cur = [(y, t)]; curx = xc
    if cur:
        cols.append(cur)
    out = []
    for col in cols:
        col.sort(key=lambda a: a[0])           # 列の中は上から下へ
        out.append("".join(t for _, t in col))
    return "".join(out)


def is_marker(word):
    c = clean_word(word[4])
    m = PAGENUM_RE.match(c)
    return bool(m and ("-" in c or "‐" in c or "－" in c))


def build_pages(doc):
    """PDF全体から {印刷ページ番号: そのページの本文（読み順）} を作る"""
    pages = {}
    layouts = []
    for idx in range(doc.page_count):
        pg = doc[idx]
        H = pg.rect.height
        words = pg.get_text("words")
        pn = page_markers(words, H)
        real = [w for w in words if not is_marker(w)]
        if len(pn) >= 2:                        # 見開き2面 → x で左右に分割
            pn.sort(key=lambda a: -a[1])
            (pr, xr), (pl, xl) = pn[0], pn[1]
            split = (xr + xl) / 2
            pages[pr] = pages.get(pr, "") + reconstruct([w for w in real if (w[0] + w[2]) / 2 > split])
            pages[pl] = pages.get(pl, "") + reconstruct([w for w in real if (w[0] + w[2]) / 2 <= split])
            layouts.append(2)
        elif len(pn) == 1:                      # 1面
            pages[pn[0][0]] = pages.get(pn[0][0], "") + reconstruct(real)
            layouts.append(1)
        else:
            layouts.append(0)
    return pages, layouts


def load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    return wb, ws, rows


def reassign(pages, rows):
    """各行を、本文ストリームに照合して印刷ページ番号を割り当てる（順番は単調）"""
    ids = sorted(pages)
    S, P = "", []
    for p in ids:
        t = norm(pages[p])
        S += t
        P += [p] * len(t)
    pos = 0
    prev = ids[0] if ids else 1
    out = []  # (row_obj, old_page, new_page, method)
    for cells in rows:
        old_page = cells[1].value
        who = cells[2].value
        text = cells[3].value
        q1 = norm((who or "") + (text or ""))
        q2 = norm(text or "")
        idx, L, m = -1, 0, "ok"
        if q1:
            j = S.find(q1, pos)
            if j >= 0:
                idx, L = j, len(q1)
        if idx < 0 and q2:
            j = S.find(q2, pos)
            if j >= 0:
                idx, L = j, len(q2)
        if idx < 0 and q1:                      # わずかな重なりだけ戻って探す
            j = S.find(q1, max(0, pos - len(q1) - 5))
            if 0 <= j < pos:
                idx, L = j, len(q1)
        if idx < 0 and q2 and len(q2) >= 6:      # ページをまたぐ／特殊文字 → 先頭で探す
            for k in (12, 10, 8, 6):
                if k > len(q2):
                    continue
                j = S.find(q2[:k], pos)
                if 0 <= j <= pos + len(q2) + 200:
                    idx, L, m = j, k, "prefix"
                    break
        if idx >= 0:
            page = P[idx]; prev = page; pos = idx + L
        else:
            page = prev; m = "carry"             # 見つからない行は直前ページに合わせる
        out.append((cells, old_page, page, m))
    return out


def main():
    if len(sys.argv) < 3:
        sys.exit("使い方: python3 fix_serifu_pages.py <台本.pdf> <AIが作った.xlsx> [出力.xlsx]")
    pdf_path, xlsx_path = sys.argv[1], sys.argv[2]
    out_path = sys.argv[3] if len(sys.argv) > 3 else re.sub(r"\.xlsx$", "", xlsx_path) + "_ページ修正版.xlsx"

    print("PDFを読み込み中 ...")
    doc = fitz.open(pdf_path)
    pages, layouts = build_pages(doc)
    if not pages:
        sys.exit("PDFからページ番号（-数字-）を見つけられませんでした。\n"
                 "各ページ下にページ番号が印刷されているPDFか確認してください。")
    twoup = layouts.count(2)
    oneup = layouts.count(1)
    print("  PDF %dページ / 検出した印刷ページ %d面（min %d, max %d）"
          % (doc.page_count, len(pages), min(pages), max(pages)))
    print("  レイアウト: 見開き2面=%d枚, 1面=%d枚, 不明=%d枚"
          % (twoup, oneup, layouts.count(0)))

    print("Excelを読み込み中 ...")
    wb, ws, rows = load_rows(xlsx_path)
    print("  %d行" % len(rows))

    print("ページを照合して振り直し中 ...")
    result = reassign(pages, rows)

    changed = sum(1 for _, o, n, _ in result if o != n)
    carried = sum(1 for _, _, _, m in result if m == "carry")
    spanning = sum(1 for _, _, _, m in result if m == "prefix")
    # 単調性チェック
    seq = [n for _, _, n, _ in result]
    backward = sum(1 for i in range(1, len(seq)) if seq[i] < seq[i - 1])

    # 書き込み
    for cells, _old, new, _m in result:
        cells[1].value = new
    wb.save(out_path)

    print("")
    print("==== 結果 ====")
    print("  ページを直した行 : %d / %d 行" % (changed, len(rows)))
    print("  ページをまたぐ行 : %d 行（始まりのページに統一）" % spanning)
    print("  照合できず前ページに合わせた行 : %d 行" % carried)
    print("  ページが逆戻りした箇所 : %d（0なら整合OK）" % backward)
    print("  保存しました → %s" % out_path)
    if backward > 0:
        print("  ※ 逆戻りがある場合は、PDFのページ番号や読み取りがうまくいっていない可能性があります。")


if __name__ == "__main__":
    main()
